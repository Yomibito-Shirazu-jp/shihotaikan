"""Print header row and 2 sample rows from each provided xlsx."""
import openpyxl, sys, json, os

paths = [
    r"C:\Users\ishij\Documents\GitHub\shihotaikan\司法大観\組版で使用したデータベース\裁001-071_最高裁判所_A00001-A00282).xlsx",
    r"C:\Users\ishij\Documents\GitHub\shihotaikan\司法大観\組版で使用したデータベース\法001-075_法務省_C00001-C000298.xlsx",
    r"C:\Users\ishij\Documents\GitHub\shihotaikan\司法大観\組版で使用したデータベース\法653-722-公証人G00001-G00277.xlsx",
]

for p in paths:
    print("=" * 80)
    print("FILE:", os.path.basename(p))
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  SHEET '{sn}'  rows={ws.max_row}  cols={ws.max_column}")
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= 3:
                break
        for i, r in enumerate(rows):
            tag = "HEAD" if i == 0 else f"r{i}"
            print(f"  {tag}: {r}")
    wb.close()
