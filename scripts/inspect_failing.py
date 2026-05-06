import openpyxl, sys
sys.stdout.reconfigure(encoding="utf-8")
wb = openpyxl.load_workbook(
    r"C:\Users\ishij\Documents\GitHub\shihotaikan\司法大観\組版で使用したデータベース\裁081-287_東京高等裁判所管内_B10001-B10828.xlsx",
    read_only=True, data_only=True,
)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
hdr = list(rows[0])
for i, h in enumerate(hdr):
    print(i, repr(h))
print("---r1 cols 19..40---")
for i in range(19, 41):
    print(i, repr(rows[1][i]))
wb.close()
