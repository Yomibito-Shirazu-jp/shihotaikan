"""Convert 司法大観 typesetting xlsx files into submissions-shaped JSON
ready for upload via /admin "過去データ入力 (JSON)".
"""
from __future__ import annotations
import openpyxl
import json
import re
import os
import glob

ROOT = os.path.join(
    r"C:\Users\ishij\Documents\GitHub\shihotaikan",
    "司法大観",                          # 司法大観
    "組版で使用したデータベース",  # 組版で使用したデータベース
)
OUT_DIR = r"C:\Users\ishij\Documents\GitHub\shihotaikan\scripts\out"
os.makedirs(OUT_DIR, exist_ok=True)

# 47 都道府県 (set membership lookup)
PREFECTURES = {
    "北海道", "青森県", "岩手県",
    "宮城県", "秋田県", "山形県",
    "福島県", "茨城県", "栃木県",
    "群馬県", "埼玉県", "千葉県",
    "東京都", "神奈川県", "新潟県",
    "富山県", "石川県", "福井県",
    "山梨県", "長野県", "岐阜県",
    "静岡県", "愛知県", "三重県",
    "滋賀県", "京都府", "大阪府",
    "兵庫県", "奈良県", "和歌山県",
    "鳥取県", "島根県", "岡山県",
    "広島県", "山口県", "徳島県",
    "香川県", "愛媛県", "高知県",
    "福岡県", "佐賀県", "長崎県",
    "熊本県", "大分県", "宮崎県",
    "鹿児島県", "沖縄県",
}

KANJI_DIGIT = {
    "〇": 0,  # 〇
    "零": 0,  # 零
    "一": 1,  # 一
    "二": 2,  # 二
    "三": 3,  # 三
    "四": 4,  # 四
    "五": 5,  # 五
    "六": 6,  # 六
    "七": 7,  # 七
    "八": 8,  # 八
    "九": 9,  # 九
    "元": 1,  # 元 (gan-nen)
}
KANJI_TEN  = "十"   # 十
KANJI_HUN  = "百"   # 百
DIGITS_CHARSET = "[〇零一二三四五六七八九十百元]"

def kanji_to_int(s: str) -> int:
    if not s:
        return 0
    if s == "元":  # 元
        return 1
    if KANJI_TEN in s or KANJI_HUN in s:
        total = 0
        cur = 0
        for ch in s:
            if ch in KANJI_DIGIT:
                cur = KANJI_DIGIT[ch]
            elif ch == KANJI_TEN:
                total += (cur if cur else 1) * 10
                cur = 0
            elif ch == KANJI_HUN:
                total += (cur if cur else 1) * 100
                cur = 0
        total += cur
        return total
    n = 0
    for ch in s:
        n = n * 10 + KANJI_DIGIT.get(ch, 0)
    return n

# 明治 大正 昭和 平成 令和
ERAS = "明治|大正|昭和|平成|令和"

BIRTH_RE = re.compile(
    "(" + ERAS + ")"
    + "(" + DIGITS_CHARSET + "+)年"   # 年
    + "(" + DIGITS_CHARSET + "+)月"   # 月
    + "(" + DIGITS_CHARSET + "+)日"   # 日
)

def parse_birthdate(s):
    if not s:
        return ("", "", "", "")
    m = BIRTH_RE.search(str(s))
    if not m:
        return ("", "", "", "")
    era, y, mo, d = m.groups()
    return (era, str(kanji_to_int(y)), str(kanji_to_int(mo)), str(kanji_to_int(d)))

# career date prefix: optional era prefix + digits 年 + digits 月 (+ optional 末日)
# 同 = "same year", 末日 = "last day"
CAREER_DATE_RE = re.compile(
    "^((?:" + ERAS + "|同)?"                                   # 同
    "(?:元|" + DIGITS_CHARSET + "+)年"                      # 元 / digits 年
    + DIGITS_CHARSET + "+月"                                    # 月
    "(?:末日)?)\\s*(.*)$"                                   # 末日
)

# Entry separators: 全角 space U+3000, Adobe-PUA paragraph mark U+E0B3,
# ASCII tab/CR/LF.
CAREER_SPLIT_RE = re.compile("[　\t\n\r]+")

def parse_career(text):
    if not text:
        return []
    parts = [p.strip() for p in CAREER_SPLIT_RE.split(str(text)) if p and p.strip()]
    out = []
    for p in parts:
        m = CAREER_DATE_RE.match(p)
        if m:
            out.append({"date": m.group(1), "content": m.group(2).strip()})
        else:
            out.append({"date": "", "content": p})
    return out

def joined_cells(row, base_idx, n=5, step=2):
    """Concatenate up to `n` cells starting at `base_idx`, jumping by `step` cols.
    The xlsx interleaves 名字漢字1, 名字読み1, 名字漢字2, 名字読み2 ... so step=2
    visits only one kind (kanji or kana) consistently."""
    chars = []
    for i in range(n):
        idx = base_idx + i * step
        if 0 <= idx < len(row):
            v = row[idx]
            if v not in (None, ""):
                chars.append(str(v))
    return "".join(chars)

# 外字 detection: anything outside ASCII printable, Hiragana, Katakana,
# CJK punctuation/symbols, CJK Unified Ideographs (incl Ext A), halfwidth/fullwidth.
GAIJI_RE = re.compile(
    "[^"
    " -~"   # ASCII printable
    "぀-ゟ"   # Hiragana
    "゠-ヿ"   # Katakana
    "　-〿"   # CJK Symbols and Punctuation
    "一-鿿"   # CJK Unified Ideographs
    "㐀-䶿"   # CJK Extension A
    "＀-￯"   # Halfwidth & Fullwidth Forms
    "]"
)

def has_gaiji(*strings) -> bool:
    for s in strings:
        if s and GAIJI_RE.search(s):
            return True
    return False

def normalize_birthplace(s):
    if not s:
        return ""
    s = str(s).strip()
    for pref in PREFECTURES:
        if pref in s:
            return pref
    return "その他"   # その他

def find_header_index(headers, name_candidates):
    for i, h in enumerate(headers):
        if h is None:
            continue
        h2 = str(h).strip()
        for cand in name_candidates:
            if h2 == cand:
                return i
    return -1

# Header names (literal-by-escape)
H_NAME    = "name"
H_PH      = "ph"
H_POST    = "post"
H_BIRTH   = "生年月日"           # 生年月日
H_ORIGIN  = "本籍"                       # 本籍
H_CAREER  = "履歴"                       # 履歴
H_TERM    = "期"                             # 期
H_LK1     = "名字漢字1"     # 名字漢字1
H_LR1     = "名字読み1"     # 名字読み1
H_FK1     = "名前漢字1"     # 名前漢字1
H_FR1     = "名前読み1"     # 名前読み1

DEPT_COURT = "裁判所の部"    # 裁判所の部
DEPT_MOJ   = "法務省の部"    # 法務省の部
PAST_LABEL = "過去データ"    # 過去データ
ZEN_SPACE  = "　"

def convert_workbook(path, dept_label):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = list(rows[0])

    idx_name      = find_header_index(headers, [H_NAME])
    idx_post      = find_header_index(headers, [H_POST])
    idx_birth     = find_header_index(headers, [H_BIRTH])
    idx_origin    = find_header_index(headers, [H_ORIGIN])
    idx_career    = find_header_index(headers, [H_CAREER])
    idx_term      = find_header_index(headers, [H_TERM])
    idx_ph        = find_header_index(headers, [H_PH])

    name_positions = [i for i, h in enumerate(headers)
                      if h is not None and str(h).strip() == H_NAME]
    idx_merged_name = name_positions[1] if len(name_positions) >= 2 else -1

    idx_lk1 = find_header_index(headers, [H_LK1])
    idx_lr1 = find_header_index(headers, [H_LR1])
    idx_fk1 = find_header_index(headers, [H_FK1])
    idx_fr1 = find_header_index(headers, [H_FR1])
    # Fallback: some workbooks have a corrupted '名字漢字1' header (e.g. 'あああ...').
    # The data is still in (位置 of 名字読み1) - 1.
    if idx_lk1 == -1 and idx_lr1 >= 1:
        idx_lk1 = idx_lr1 - 1
    if idx_fk1 == -1 and idx_fr1 >= 1:
        idx_fk1 = idx_fr1 - 1

    out = []
    src_basename = os.path.basename(path)
    for r in rows[1:]:
        if not r or all(v in (None, "") for v in r):
            continue
        name_kanji_full = r[idx_name] if idx_name >= 0 else None
        post = r[idx_post] if idx_post >= 0 else None
        if not name_kanji_full and not post:
            continue

        last_name  = joined_cells(r, idx_lk1) if idx_lk1 >= 0 else ""
        first_name = joined_cells(r, idx_fk1) if idx_fk1 >= 0 else ""
        last_kana  = joined_cells(r, idx_lr1) if idx_lr1 >= 0 else ""
        first_kana = joined_cells(r, idx_fr1) if idx_fr1 >= 0 else ""

        if (not last_name or not first_name) and name_kanji_full:
            cleaned = re.sub(ZEN_SPACE + "+", "", str(name_kanji_full)).strip()
            if not last_name and not first_name:
                last_name = cleaned

        merged_kanji = ""
        if idx_merged_name >= 0:
            v = r[idx_merged_name]
            if v not in (None, ""):
                merged_kanji = str(v).strip()

        birth_era, birth_y, birth_m, birth_d = parse_birthdate(
            r[idx_birth] if idx_birth >= 0 else ""
        )
        career = parse_career(r[idx_career] if idx_career >= 0 else "")
        birthplace = normalize_birthplace(r[idx_origin] if idx_origin >= 0 else "")

        term_val = r[idx_term] if idx_term >= 0 else None
        try:
            training_term = str(int(term_val)) if term_val not in (None, "", "-") else ""
        except (TypeError, ValueError):
            training_term = ""

        ph_code = str(r[idx_ph] if idx_ph >= 0 else "").strip()
        gaiji = has_gaiji(last_name, first_name,
                          name_kanji_full or "", merged_kanji)
        # Storage path mirrors the source: past_photos/{prefix}/{ph_code}.jpg
        # where prefix is the first character of the ph code (A/B/C/D/E/F/G).
        # The 司法大観/画像/Links/{prefix}/ folders use the same scheme.
        if ph_code:
            prefix = ph_code[0].upper()
            past_photo_path = f"past_photos/{prefix}/{ph_code}.jpg"
        else:
            past_photo_path = ""

        rec = {
            "department": dept_label,
            "lastName": last_name,
            "firstName": first_name,
            "lastNameKana": last_kana,
            "firstNameKana": first_kana,
            "jobTitle": str(post or "").strip(),
            "mergedLastName": "",
            "birthEra": birth_era,
            "birthYear": birth_y,
            "birthMonth": birth_m,
            "birthDay": birth_d,
            "birthPlace": birthplace,
            "careerType": PAST_LABEL,
            "careerNew": career,
            "oldCareerDepartment": dept_label,
            "oldCareerPage": "",
            "careerAdd": [],
            "photoType": PAST_LABEL,
            "oldPhotoDepartment": dept_label,
            "oldPhotoPage": "",
            "trainingTerm": training_term,
            "agreeTerms": True,
            "needsGarbledTextCheck": gaiji,
            "userId": "admin_import",
            "pastPhotoCode": ph_code,
            "pastPhotoStoragePath": past_photo_path,
            "_sourceFile": src_basename,
        }
        out.append(rec)
    return out


def main():
    files = sorted(glob.glob(os.path.join(ROOT, "*.xlsx")))
    all_records = []
    per_file_counts = []
    for p in files:
        bn = os.path.basename(p)
        if bn.startswith("裁"):       # 裁
            dept = DEPT_COURT
        elif bn.startswith("法"):     # 法
            dept = DEPT_MOJ
        else:
            dept = ""
        recs = convert_workbook(p, dept)
        per_file_counts.append((bn, len(recs)))
        all_records.extend(recs)

    out_path = os.path.join(OUT_DIR, "past_submissions.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)

    print("=" * 72)
    for bn, c in per_file_counts:
        print(f"{c:5d}  {bn}")
    print("-" * 72)
    print(f"TOTAL records: {len(all_records)}")
    gaiji_n = sum(1 for r in all_records if r["needsGarbledTextCheck"])
    print(f"  with gaiji flag: {gaiji_n}")
    print(f"Output: {out_path}")
    if all_records:
        keep_keys = {"department","lastName","firstName","lastNameKana",
                     "firstNameKana","jobTitle","birthEra","birthYear",
                     "birthMonth","birthDay","birthPlace","trainingTerm",
                     "pastPhotoCode","pastPhotoStoragePath"}
        sample = {k: v for k, v in all_records[0].items() if k in keep_keys}
        print("Sample[0]:", json.dumps(sample, ensure_ascii=False))
        print("Sample[0].careerNew[:3]:",
              json.dumps(all_records[0]["careerNew"][:3], ensure_ascii=False))

if __name__ == "__main__":
    main()
